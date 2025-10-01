import typer
from typing import List, Optional
from pathlib import Path
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
from rich import print as rprint
from collections import defaultdict, Counter
import pydicom

# Excel availability check
try:
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import PieChart, BarChart, Reference
    from openpyxl.chart.series import DataPoint, SeriesLabel
    from openpyxl.utils.dataframe import dataframe_to_rows
    EXCEL_AVAILABLE = True
except ImportError as e:
    EXCEL_AVAILABLE = False
    pd = None
    console = Console()  # Create console here for error message
    console.print(f"⚠️  Excel dependencies not available: {e}", style="yellow")

from dicom_compare.dicom_extractor import DicomExtractor, ExtractionStats
from dicom_compare.dicom_loader import DicomLoader
from dicom_compare.dicom_comparator import DicomComparator
from dicom_compare.models import ComparisonSummary, FileComparisonResult
from dicom_compare.utils import validate_inputs, create_temp_dir, cleanup_temp_dirs
from dicom_compare.image_command import run_image_comparison
from dicom_compare.hierarchical_loader import HierarchicalDicomLoader
from dicom_compare.tag_search import TagSearchEngine, InteractiveSearchSession

app = typer.Typer(
    name="dicomcompare",
    help="Compare DICOM studies from different ZIP exports to identify differences",
    add_completion=False
)

console = Console()

@app.command("image")
def compare_images(
    files: List[Path] = typer.Option(
        ..., 
        "-f", 
        "--file", 
        help="ZIP files to compare (first file is baseline, minimum 2 files required)"
    ),
    report: Optional[Path] = typer.Option(
        None,
        "-r",
        "--report",
        help="Path to save image comparison report (CSV/Excel format)"
    ),
    tolerance: float = typer.Option(
        0.0,
        "-t",
        "--tolerance",
        help="Tolerance for pixel differences (0.0 = exact match, higher = more tolerant)"
    ),
    normalize: bool = typer.Option(
        True,
        "--normalize/--no-normalize",
        help="Apply DICOM normalization (rescale slope/intercept, window/level)"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Compare DICOM image pixel data between studies.
    
    This command compares the actual pixel values of DICOM images rather than 
    just the metadata tags. Useful for validating that image data is preserved 
    across different export methods.
    """
    run_image_comparison(files, report, tolerance, normalize, verbose)

# Create inspect command group
inspect_app = typer.Typer(
    name="inspect",
    help="Inspect DICOM content at different hierarchy levels"
)
app.add_typer(inspect_app, name="inspect")

@inspect_app.command("files")
def inspect_files(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to inspect"
    )
):
    """
    Inspect ZIP file structure and basic DICOM content (original functionality).
    """
    console.print("🔍 Inspecting ZIP files...", style="blue")

    temp_dirs = []
    try:
        extractor = DicomExtractor()

        for file in files:
            console.print(f"\n📦 Inspecting {file.name}:", style="bold cyan")

            temp_dir = create_temp_dir()
            temp_dirs.append(temp_dir)

            # Extract
            extracted_path, stats = extractor.extract_zip(file, temp_dir)

            # Find DICOMs
            dicom_files = extractor.find_dicom_files(extracted_path)

            if dicom_files:
                console.print(f"\n✅ Found {len(dicom_files)} DICOM files", style="green")

                # Group by directory
                by_directory = defaultdict(list)
                for dicom_file in dicom_files:
                    relative_path = dicom_file.relative_to(extracted_path)
                    directory = str(relative_path.parent)
                    by_directory[directory].append(relative_path.name)

                for directory, files in by_directory.items():
                    console.print(f"   📁 {directory}: {len(files)} DICOM files", style="cyan")
                    for dicom_file in files[:3]:  # Show first 3 files per directory
                        console.print(f"      {dicom_file}", style="dim")
                    if len(files) > 3:
                        console.print(f"      ... and {len(files) - 3} more files", style="dim")

                # Load first few to check SOPInstanceUIDs
                console.print(f"\n🔍 Checking DICOM content:", style="cyan")
                for i, dicom_file in enumerate(dicom_files[:5]):  # Check first 5
                    try:

                        ds = pydicom.dcmread(dicom_file, stop_before_pixels=True)
                        sop_uid = getattr(ds, 'SOPInstanceUID', 'MISSING')
                        series_uid = getattr(ds, 'SeriesInstanceUID', 'MISSING')
                        relative_path = dicom_file.relative_to(extracted_path)
                        console.print(f"   📄 {relative_path}", style="dim")
                        console.print(f"      SOPInstanceUID = {sop_uid}", style="dim")
                        console.print(f"      SeriesInstanceUID = {series_uid}", style="dim")
                    except Exception as e:
                        console.print(f"   ❌ {dicom_file.name}: Error reading - {e}", style="red")

                if len(dicom_files) > 5:
                    console.print(f"   ... and {len(dicom_files) - 5} more DICOM files", style="dim")
            else:
                console.print("❌ No DICOM files found", style="red")

    finally:
        cleanup_temp_dirs(temp_dirs)

@inspect_app.command("search")
def inspect_search(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to search"
    ),
    interactive: bool = typer.Option(
        True,
        "--interactive/--no-interactive",
        help="Start interactive search session"
    ),
    max_results: int = typer.Option(
        20,
        "--max-results",
        help="Maximum number of search results"
    ),
    query: Optional[str] = typer.Option(
        None,
        "-q",
        "--query",
        help="Search query (if not provided, starts interactive mode)"
    )
):
    """
    Interactive fuzzy search across all DICOM tags.
    """
    console.print("🔍 Loading DICOM files for search...", style="blue")

    try:
        # Load hierarchical data
        loader = HierarchicalDicomLoader(verbose=False)  # Search is always quiet by default
        data = loader.load_hierarchical_data(files)

        # Create search engine
        search_engine = TagSearchEngine(data)

        # Show loading results
        stats = data.get_stats()
        console.print(f"📊 Loaded: {stats['patients']} patients, {stats['studies']} studies, "
                     f"{stats['series']} series, {stats['instances']} instances", style="green")

        if query and not interactive:
            # Non-interactive single query
            results = search_engine.fuzzy_search(query, max_results=max_results)
            _display_search_results_brief(results, query, console)
        else:
            # Interactive search session
            session = InteractiveSearchSession(search_engine, console)
            session.start_session()

    except Exception as e:
        console.print(f"❌ Search failed: {str(e)}", style="red")
        raise typer.Exit(1)

@inspect_app.command("patient")
def inspect_patient(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to inspect"
    ),
    patient_id: Optional[str] = typer.Option(
        None,
        "--patient-id",
        help="Specific patient ID to inspect"
    ),
    anonymize: bool = typer.Option(
        False,
        "--anonymize",
        help="Anonymize patient identifiers"
    ),
    show_studies: bool = typer.Option(
        True,
        "--studies/--no-studies",
        help="Show patient studies summary"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Inspect patient-level DICOM tags and demographics.
    """
    console.print("👤 Inspecting patient information...", style="blue")

    try:
        # Load hierarchical data
        loader = HierarchicalDicomLoader(verbose=verbose)
        data = loader.load_hierarchical_data(files)

        # Filter by patient ID if specified
        patients_to_show = {}
        if patient_id:
            # Look for exact or fuzzy match
            found_patient = None
            for pid, patient in data.patients.items():
                if pid == patient_id or patient_id.lower() in pid.lower():
                    found_patient = patient
                    patients_to_show[pid] = patient
                    break

            if not found_patient:
                console.print(f"❌ Patient ID '{patient_id}' not found", style="red")
                _list_available_patients(data.patients, console)
                return
        else:
            patients_to_show = data.patients

        # Display patient information
        _display_patient_info(patients_to_show, data, anonymize, show_studies, console)

    except Exception as e:
        console.print(f"❌ Patient inspection failed: {str(e)}", style="red")
        raise typer.Exit(1)

@inspect_app.command("study")
def inspect_study(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to inspect"
    ),
    study_uid: Optional[str] = typer.Option(
        None,
        "--study-uid",
        help="Specific study UID to inspect"
    ),
    patient_id: Optional[str] = typer.Option(
        None,
        "--patient-id",
        help="Filter by patient ID"
    ),
    show_series: bool = typer.Option(
        True,
        "--series/--no-series",
        help="Show study series overview"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Inspect study-level DICOM tags and acquisition details.
    """
    console.print("📚 Inspecting study information...", style="blue")

    try:
        # Load hierarchical data
        loader = HierarchicalDicomLoader(verbose=verbose)
        data = loader.load_hierarchical_data(files)

        # Filter studies
        studies_to_show = {}

        if study_uid:
            # Look for specific study
            found_study = None
            for uid, study in data.studies.items():
                if uid == study_uid or study_uid in uid:
                    found_study = study
                    studies_to_show[uid] = study
                    break

            if not found_study:
                console.print(f"❌ Study UID '{study_uid}' not found", style="red")
                _list_available_studies(data.studies, console)
                return
        elif patient_id:
            # Filter by patient
            patient_found = False
            for uid, study in data.studies.items():
                if study.patient_id == patient_id or patient_id.lower() in study.patient_id.lower():
                    studies_to_show[uid] = study
                    patient_found = True

            if not patient_found:
                console.print(f"❌ No studies found for patient '{patient_id}'", style="red")
                return
        else:
            studies_to_show = data.studies

        # Display study information
        _display_study_info(studies_to_show, data, show_series, console)

    except Exception as e:
        console.print(f"❌ Study inspection failed: {str(e)}", style="red")
        raise typer.Exit(1)

@inspect_app.command("series")
def inspect_series(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to inspect"
    ),
    series_uid: Optional[str] = typer.Option(
        None,
        "--series-uid",
        help="Specific series UID to inspect"
    ),
    study_uid: Optional[str] = typer.Option(
        None,
        "--study-uid",
        help="Filter by study UID"
    ),
    show_instances: bool = typer.Option(
        True,
        "--instances/--no-instances",
        help="Show series instances overview"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Inspect series-level DICOM tags and acquisition parameters.
    """
    console.print("🔬 Inspecting series information...", style="blue")

    try:
        # Load hierarchical data
        loader = HierarchicalDicomLoader(verbose=verbose)
        data = loader.load_hierarchical_data(files)

        # Filter series
        series_to_show = {}

        if series_uid:
            # Look for specific series
            found_series = None
            for uid, series in data.series.items():
                if uid == series_uid or series_uid in uid:
                    found_series = series
                    series_to_show[uid] = series
                    break

            if not found_series:
                console.print(f"❌ Series UID '{series_uid}' not found", style="red")
                _list_available_series(data.series, console)
                return
        elif study_uid:
            # Filter by study
            study_found = False
            for uid, series in data.series.items():
                if series.study_uid == study_uid or study_uid in series.study_uid:
                    series_to_show[uid] = series
                    study_found = True

            if not study_found:
                console.print(f"❌ No series found for study '{study_uid}'", style="red")
                return
        else:
            series_to_show = data.series

        # Display series information
        _display_series_info(series_to_show, data, show_instances, console)

    except Exception as e:
        console.print(f"❌ Series inspection failed: {str(e)}", style="red")
        raise typer.Exit(1)

@inspect_app.command("instance")
def inspect_instance(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to inspect"
    ),
    sop_uid: Optional[str] = typer.Option(
        None,
        "--sop-uid",
        help="Specific SOP instance UID to inspect"
    ),
    series_uid: Optional[str] = typer.Option(
        None,
        "--series-uid",
        help="Filter by series UID"
    ),
    show_all_tags: bool = typer.Option(
        False,
        "--all-tags",
        help="Show all DICOM tags (not just key ones)"
    ),
    limit: int = typer.Option(
        10,
        "--limit",
        help="Maximum number of instances to display"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Inspect instance-level DICOM tags and technical parameters.
    """
    console.print("🖼️  Inspecting instance information...", style="blue")

    try:
        # Load hierarchical data
        loader = HierarchicalDicomLoader(verbose=verbose)
        data = loader.load_hierarchical_data(files)

        # Filter instances
        instances_to_show = {}

        if sop_uid:
            # Look for specific instance
            found_instance = None
            for uid, instance in data.instances.items():
                if uid == sop_uid or sop_uid in uid:
                    found_instance = instance
                    instances_to_show[uid] = instance
                    break

            if not found_instance:
                console.print(f"❌ SOP Instance UID '{sop_uid}' not found", style="red")
                _list_available_instances(data.instances, console, limit=5)
                return
        elif series_uid:
            # Filter by series
            series_found = False
            for uid, instance in data.instances.items():
                if instance.series_uid == series_uid or series_uid in instance.series_uid:
                    instances_to_show[uid] = instance
                    series_found = True

                    # Respect limit
                    if len(instances_to_show) >= limit:
                        break

            if not series_found:
                console.print(f"❌ No instances found for series '{series_uid}'", style="red")
                return
        else:
            # Show first N instances
            instances_to_show = dict(list(data.instances.items())[:limit])

        # Display instance information
        _display_instance_info(instances_to_show, data, show_all_tags, console)

    except Exception as e:
        console.print(f"❌ Instance inspection failed: {str(e)}", style="red")
        raise typer.Exit(1)

@app.command()
def compare(
    files: List[Path] = typer.Option(
        ...,
        "-f",
        "--file",
        help="ZIP files to compare (first file is baseline, minimum 2 files required)"
    ),
    report: Optional[Path] = typer.Option(
        None,
        "-r",
        "--report",
        help="Path to save CSV/Excel report (format determined by extension)"
    ),
    matching_mode: str = typer.Option(
        "uid",
        "--matching-mode",
        help="Matching strategy: 'uid' (default), 'hash' (pixel hash), 'fingerprint' (statistical), 'smart' (cascading fallback)"
    ),
    verbose: bool = typer.Option(
        False,
        "-v",
        "--verbose",
        help="Enable verbose output"
    )
):
    """
    Compare DICOM studies from ZIP files.
    
    The first file specified with -f is treated as the baseline/original.
    All subsequent files are compared against this baseline.
    """
    
    # Validate inputs
    try:
        validate_inputs(files)
        if report:
            validate_report_path(report)

        # Validate matching mode
        valid_modes = ["uid", "hash", "fingerprint", "smart"]
        if matching_mode not in valid_modes:
            raise ValueError(f"Invalid matching mode '{matching_mode}'. Must be one of: {', '.join(valid_modes)}")

    except Exception as e:
        console.print(f"❌ {str(e)}", style="red")
        raise typer.Exit(1)
    
    console.print("🔍 Starting DICOM comparison...", style="blue")
    
    temp_dirs = []
    extraction_stats = []
    try:
        # Extract ZIP files
        console.print("📦 Extracting ZIP files...", style="yellow")
        extractor = DicomExtractor(verbose=verbose)
        extracted_paths = []
        
        for file in files:
            temp_dir = create_temp_dir()
            temp_dirs.append(temp_dir)
            extracted_path, stats = extractor.extract_zip(file, temp_dir)
            extracted_paths.append((str(file), extracted_path))
            extraction_stats.append((str(file), stats))
        
        # Load DICOM files
        console.print("🏥 Loading DICOM files...", style="yellow")
        loader = DicomLoader(verbose=verbose)
        loaded_studies = []
        
        for i, (file_name, path) in enumerate(extracted_paths):
            studies = loader.load_dicom_files(path, file_name)
            loaded_studies.append((file_name, studies))
            
            # Show results with extraction context
            total_instances = sum(len(series.instances) for study in studies.values() 
                                for series in study.series.values())
            
            # Get corresponding extraction stats
            _, stats = extraction_stats[i]
            
            if stats.non_dicom_files > 0:
                console.print(f"   {Path(file_name).name}: {total_instances} instances ({stats.dicom_files}/{stats.total_files} files were DICOM)", style="cyan")
            else:
                console.print(f"   {Path(file_name).name}: {total_instances} instances (all {stats.total_files} files were DICOM)", style="cyan")
            
            if verbose:
                console.print(f"     Folders: {stats.total_folders}, DICOM files: {stats.dicom_files}, Skipped: {stats.non_dicom_files}", style="dim")
    
        
        # Compare studies
        console.print(f"🔍 Comparing DICOM studies (matching mode: {matching_mode})...", style="yellow")
        comparator = DicomComparator()

        baseline_name, baseline_studies = loaded_studies[0]
        comparison_results = []

        for comp_name, comp_studies in loaded_studies[1:]:
            result = comparator.compare_studies(
                baseline_studies, comp_studies,
                baseline_name, comp_name,
                matching_mode=matching_mode
            )
            comparison_results.append(result)
        
        if verbose:
            console.print("\n🔍 Comparison Debug Info:", style="cyan")
            for i, (comp_name, comp_studies) in enumerate(loaded_studies[1:]):
                result = comparison_results[i]
                console.print(f"   {Path(comp_name).name}:", style="cyan")
                console.print(f"     Baseline instances: {result.total_instances_baseline}", style="dim")
                console.print(f"     Comparison instances: {result.total_instances_comparison}", style="dim")
                console.print(f"     Matched instances: {len(result.matched_instances)}", style="dim")
                console.print(f"     Perfect matches: {sum(1 for comp in result.matched_instances if comp.is_perfect_match)}", style="dim")
                console.print(f"     Tag differences: {sum(1 for comp in result.matched_instances if not comp.is_perfect_match)}", style="dim")
                console.print(f"     Missing instances: {len(result.missing_instances)}", style="dim")
                console.print(f"     Extra instances: {len(result.extra_instances)}", style="dim")
        
        # Create summary
        summary = create_comparison_summary(baseline_name, comparison_results)
        
        # Display results
        display_terminal_results(summary, console)
        
        # Generate report if requested
        if report:
            console.print(f"📊 Generating report: {report}", style="green")
            generate_report(summary, report)
            console.print(f"✅ Report saved to: {report}", style="green")
        
        console.print("🎉 Comparison completed successfully!", style="green")
        
    except Exception as e:
        console.print(f"❌ Error during comparison: {str(e)}", style="red")
        if verbose:
            console.print_exception()
        raise typer.Exit(1)
    
    finally:
        # Cleanup temporary directories
        cleanup_temp_dirs(temp_dirs)

def validate_report_path(report_path: Path) -> None:
    """Validate report path and format"""
    if not report_path.suffix.lower() in ['.csv', '.xlsx']:
        raise ValueError("Report format must be CSV (.csv) or Excel (.xlsx)")
    
    # Ensure parent directory exists
    report_path.parent.mkdir(parents=True, exist_ok=True)

def create_comparison_summary(baseline_name: str, results: List[FileComparisonResult]) -> ComparisonSummary:
    """Create summary from comparison results"""
    comparison_files = [result.comparison_file for result in results]
    
    # Calculate totals from results
    total_instances = results[0].total_instances_baseline if results else 0
    
    # Count unique studies and series from baseline
    total_studies = 1  # Simplified - we'll improve this later
    total_series = 1   # Simplified - we'll improve this later
    
    return ComparisonSummary(
        baseline_file=baseline_name,
        comparison_files=comparison_files,
        file_results=results,
        total_instances=total_instances,
        total_studies=total_studies,
        total_series=total_series
    )

def display_terminal_results(summary: 'ComparisonSummary', console: Console) -> None:
    """Display formatted results in terminal"""
    # Create summary panel
    summary_text = f"Baseline: {Path(summary.baseline_file).name}\n"
    summary_text += f"Comparisons: {len(summary.comparison_files)}\n"
    summary_text += f"Total Studies: {summary.total_studies}\n"
    summary_text += f"Total Instances: {summary.total_instances}"
    
    console.print(Panel(summary_text, title="📋 DICOM Comparison Summary", expand=False))
    
    # Create comprehensive results table
    table = Table(title="🔍 Detailed Comparison Results")
    table.add_column("File", style="cyan", width=25)
    table.add_column("Perfect\nMatches", style="green", justify="right")
    table.add_column("Tag\nDiffs", style="yellow", justify="right")
    table.add_column("Tag Diff\n%", style="bright_yellow", justify="right")
    table.add_column("Missing\nInstances", style="red", justify="right")
    table.add_column("Missing\n%", style="bright_red", justify="right")  # New
    table.add_column("Extra\nInstances", style="magenta", justify="right")
    table.add_column("Extra\n%", style="bright_magenta", justify="right")  # New
    table.add_column("Data\nIntegrity", style="bright_blue", justify="right")  # New
    
    for result in summary.file_results:
        perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
        tag_diffs = len(result.matched_instances) - perfect_matches
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        
        total_baseline = result.total_instances_baseline
        total_comparison = result.total_instances_comparison
        
        # Calculate percentages
        tag_diff_pct = (tag_diffs / total_baseline * 100) if total_baseline > 0 else 0
        missing_pct = (missing / total_baseline * 100) if total_baseline > 0 else 0
        extra_pct = (extra / total_comparison * 100) if total_comparison > 0 else 0
        
        # Calculate data integrity score (0-100%)
        integrity_score = _calculate_data_integrity(result)
        
        # Color code integrity score
        if integrity_score >= 95:
            integrity_style = "bright_green"
        elif integrity_score >= 85:
            integrity_style = "green"
        elif integrity_score >= 70:
            integrity_style = "yellow"
        else:
            integrity_style = "red"
        
        table.add_row(
            Path(result.comparison_file).name,
            str(perfect_matches),
            str(tag_diffs),
            f"{tag_diff_pct:.1f}%",
            str(missing),
            f"{missing_pct:.1f}%",
            str(extra),
            f"{extra_pct:.1f}%",
            f"[{integrity_style}]{integrity_score:.1f}%[/{integrity_style}]"
        )
    
    console.print(table)
    
    # Add detailed breakdown
    _display_detailed_breakdown(summary, console)
    
    # Add tag analysis
    _display_tag_analysis(summary, console)

def _calculate_data_integrity(result: 'FileComparisonResult') -> float:
    """Calculate overall data integrity score (0-100%)"""
    total_baseline = result.total_instances_baseline
    if total_baseline == 0:
        return 0.0
    
    # Perfect matches get full score
    perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
    perfect_score = (perfect_matches / total_baseline) * 100
    
    # Tag differences get partial score (75% of full score)
    tag_diffs = len(result.matched_instances) - perfect_matches
    partial_score = (tag_diffs / total_baseline) * 75
    
    # Missing instances get no score
    # Extra instances don't affect score (they're bonus data)
    
    return perfect_score + partial_score

def _display_detailed_breakdown(summary: 'ComparisonSummary', console: Console) -> None:
    """Display detailed breakdown of differences"""
    console.print("\n")
    
    breakdown_table = Table(title="📊 Export Quality Breakdown", show_header=True)
    breakdown_table.add_column("File", style="cyan")
    breakdown_table.add_column("Instance\nMatch Rate", style="green", justify="right")
    breakdown_table.add_column("Tag\nPreservation", style="blue", justify="right")
    breakdown_table.add_column("Quality\nGrade", style="bright_white", justify="center")
    breakdown_table.add_column("Primary Issues", style="yellow")
    
    for result in summary.file_results:
        total_baseline = result.total_instances_baseline
        total_comparison = result.total_instances_comparison
        matched_instances = len(result.matched_instances)
        
        # Instance match rate
        instance_match_rate = (matched_instances / total_baseline * 100) if total_baseline > 0 else 0
        
        # Tag preservation rate (for matched instances)
        if matched_instances > 0:
            perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
            tag_preservation = (perfect_matches / matched_instances * 100)
        else:
            tag_preservation = 0
        
        # Quality grade
        integrity_score = _calculate_data_integrity(result)
        if integrity_score >= 95:
            grade = "A+"
        elif integrity_score >= 90:
            grade = "A"
        elif integrity_score >= 85:
            grade = "B+"
        elif integrity_score >= 80:
            grade = "B"
        elif integrity_score >= 70:
            grade = "C"
        elif integrity_score >= 60:
            grade = "D"
        else:
            grade = "F"
        
        # Identify primary issues
        issues = []
        if len(result.missing_instances) > total_baseline * 0.05:  # >5% missing
            issues.append(f"{len(result.missing_instances)} missing instances")
        if len(result.extra_instances) > total_comparison * 0.05:  # >5% extra
            issues.append(f"{len(result.extra_instances)} extra instances")
        
        tag_diffs = matched_instances - sum(1 for comp in result.matched_instances if comp.is_perfect_match)
        if tag_diffs > matched_instances * 0.1:  # >10% have tag differences
            issues.append(f"{tag_diffs} instances with tag changes")
        
        if not issues:
            issues.append("None detected")
        
        breakdown_table.add_row(
            Path(result.comparison_file).name,
            f"{instance_match_rate:.1f}%",
            f"{tag_preservation:.1f}%",
            grade,
            ", ".join(issues[:2])  # Show first 2 issues
        )
    
    console.print(breakdown_table)

def _display_tag_analysis(summary: 'ComparisonSummary', console: Console) -> None:
    """Display tag-level analysis"""
    
    # Collect tag difference statistics
    tag_stats = defaultdict(lambda: {'missing': 0, 'extra': 0, 'value_diff': 0, 'type_diff': 0})
    total_differences = 0
    
    for result in summary.file_results:
        for instance_comp in result.matched_instances:
            if not instance_comp.is_perfect_match:
                for tag_diff in instance_comp.tag_differences:
                    tag_name = tag_diff.tag_keyword
                    diff_type = tag_diff.difference_type.value
                    
                    if diff_type == 'MISSING_TAG':
                        tag_stats[tag_name]['missing'] += 1
                    elif diff_type == 'EXTRA_TAG':
                        tag_stats[tag_name]['extra'] += 1
                    elif diff_type == 'VALUE_DIFF':
                        tag_stats[tag_name]['value_diff'] += 1
                    elif diff_type == 'TYPE_DIFF':
                        tag_stats[tag_name]['type_diff'] += 1
                    
                    total_differences += 1
    
    if tag_stats:
        console.print("\n")
        
        # Tag differences table
        tag_table = Table(title="🏷️ Tag Difference Analysis", show_header=True)
        tag_table.add_column("Tag", style="cyan")
        tag_table.add_column("Missing", style="red", justify="right")
        tag_table.add_column("Extra", style="magenta", justify="right")
        tag_table.add_column("Value\nChanged", style="yellow", justify="right")
        tag_table.add_column("Type\nChanged", style="orange3", justify="right")
        tag_table.add_column("Total\nAffected", style="bright_white", justify="right")
        tag_table.add_column("% of\nDifferences", style="bright_blue", justify="right")
        
        # Sort by total impact
        sorted_tags = sorted(tag_stats.items(), 
                           key=lambda x: sum(x[1].values()), 
                           reverse=True)
        
        for tag_name, stats in sorted_tags[:15]:  # Show top 15
            total_tag_diffs = sum(stats.values())
            diff_percentage = (total_tag_diffs / total_differences * 100) if total_differences > 0 else 0
            
            tag_table.add_row(
                tag_name,
                str(stats['missing']) if stats['missing'] > 0 else "-",
                str(stats['extra']) if stats['extra'] > 0 else "-",
                str(stats['value_diff']) if stats['value_diff'] > 0 else "-",
                str(stats['type_diff']) if stats['type_diff'] > 0 else "-",
                str(total_tag_diffs),
                f"{diff_percentage:.1f}%"
            )
        
        console.print(tag_table)
        
        if len(sorted_tags) > 15:
            console.print(f"   ... and {len(sorted_tags) - 15} more tags with differences", style="dim")
        
        # Summary of difference types
        console.print("\n")
        _display_difference_type_summary(tag_stats, total_differences, console)

def _display_difference_type_summary(tag_stats: dict, total_differences: int, console: Console) -> None:
    """Display summary of difference types"""
    # Count by difference type
    type_counts = {'missing': 0, 'extra': 0, 'value_diff': 0, 'type_diff': 0}
    
    for stats in tag_stats.values():
        for diff_type, count in stats.items():
            type_counts[diff_type] += count
    
    summary_table = Table(title="📈 Difference Type Summary", show_header=True)
    summary_table.add_column("Difference Type", style="cyan")
    summary_table.add_column("Count", style="bright_white", justify="right")
    summary_table.add_column("Percentage", style="bright_blue", justify="right")
    summary_table.add_column("Impact", style="yellow")
    
    # Define impact descriptions
    impact_descriptions = {
        'missing': "Tags removed during export",
        'extra': "Tags added during export", 
        'value_diff': "Tag values modified",
        'type_diff': "Tag data types changed"
    }
    
    for diff_type, count in type_counts.items():
        if count > 0:
            percentage = (count / total_differences * 100) if total_differences > 0 else 0
            summary_table.add_row(
                diff_type.replace('_', ' ').title(),
                str(count),
                f"{percentage:.1f}%",
                impact_descriptions[diff_type]
            )
    
    console.print(summary_table)

def generate_report(summary: ComparisonSummary, report_path: Path) -> None:
    """Generate CSV or Excel report"""
    if report_path.suffix.lower() == '.csv':
        generate_csv_report(summary, report_path)
    elif report_path.suffix.lower() == '.xlsx':
        generate_excel_report(summary, report_path)

def generate_csv_report(summary: ComparisonSummary, report_path: Path) -> None:
    """Generate CSV report"""    
    rows = []
    
    # Add summary information first
    summary_rows = []
    for result in summary.file_results:
        perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
        tag_diffs = len(result.matched_instances) - perfect_matches
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        
        summary_rows.append({
            'ReportType': 'SUMMARY',
            'BaselineFile': Path(result.baseline_file).name,
            'ComparisonFile': Path(result.comparison_file).name,
            'SOPInstanceUID': 'SUMMARY',
            'TagName': 'TotalInstances',
            'TagKeyword': 'TotalInstances',
            'BaselineValue': str(result.total_instances_baseline),
            'ComparisonValue': str(result.total_instances_comparison),
            'DifferenceType': 'SUMMARY',
            'VR': 'SUMMARY'
        })
        
        summary_rows.append({
            'ReportType': 'SUMMARY',
            'BaselineFile': Path(result.baseline_file).name,
            'ComparisonFile': Path(result.comparison_file).name,
            'SOPInstanceUID': 'SUMMARY',
            'TagName': 'PerfectMatches',
            'TagKeyword': 'PerfectMatches',
            'BaselineValue': str(perfect_matches),
            'ComparisonValue': str(perfect_matches),
            'DifferenceType': 'SUMMARY',
            'VR': 'SUMMARY'
        })
        
        summary_rows.append({
            'ReportType': 'SUMMARY',
            'BaselineFile': Path(result.baseline_file).name,
            'ComparisonFile': Path(result.comparison_file).name,
            'SOPInstanceUID': 'SUMMARY',
            'TagName': 'TagDifferences',
            'TagKeyword': 'TagDifferences',
            'BaselineValue': str(tag_diffs),
            'ComparisonValue': str(tag_diffs),
            'DifferenceType': 'SUMMARY',
            'VR': 'SUMMARY'
        })
    
    rows.extend(summary_rows)
    
    # Add detailed differences
    difference_count = 0
    for result in summary.file_results:
        # Add missing instances
        for missing_instance in result.missing_instances:
            rows.append({
                'ReportType': 'MISSING_INSTANCE',
                'BaselineFile': Path(result.baseline_file).name,
                'ComparisonFile': Path(result.comparison_file).name,
                'SOPInstanceUID': missing_instance.sop_instance_uid,
                'TagName': 'MISSING_INSTANCE',
                'TagKeyword': 'MISSING_INSTANCE',
                'BaselineValue': 'EXISTS',
                'ComparisonValue': 'MISSING',
                'DifferenceType': 'MISSING_INSTANCE',
                'VR': 'INSTANCE'
            })
            difference_count += 1
        
        # Add extra instances
        for extra_instance in result.extra_instances:
            rows.append({
                'ReportType': 'EXTRA_INSTANCE',
                'BaselineFile': Path(result.baseline_file).name,
                'ComparisonFile': Path(result.comparison_file).name,
                'SOPInstanceUID': extra_instance.sop_instance_uid,
                'TagName': 'EXTRA_INSTANCE',
                'TagKeyword': 'EXTRA_INSTANCE',
                'BaselineValue': 'MISSING',
                'ComparisonValue': 'EXISTS',
                'DifferenceType': 'EXTRA_INSTANCE',
                'VR': 'INSTANCE'
            })
            difference_count += 1
        
        # Add tag differences
        for instance_comp in result.matched_instances:
            if not instance_comp.is_perfect_match:
                for tag_diff in instance_comp.tag_differences:
                    rows.append({
                        'ReportType': 'TAG_DIFFERENCE',
                        'BaselineFile': Path(result.baseline_file).name,
                        'ComparisonFile': Path(result.comparison_file).name,
                        'SOPInstanceUID': instance_comp.sop_instance_uid,
                        'TagName': tag_diff.tag_name,
                        'TagKeyword': tag_diff.tag_keyword,
                        'BaselineValue': str(tag_diff.baseline_value) if tag_diff.baseline_value is not None else 'NULL',
                        'ComparisonValue': str(tag_diff.comparison_value) if tag_diff.comparison_value is not None else 'NULL',
                        'DifferenceType': tag_diff.difference_type.value,
                        'VR': tag_diff.vr
                    })
                    difference_count += 1
    
    # If no differences found, add a note
    if difference_count == 0:
        rows.append({
            'ReportType': 'INFO',
            'BaselineFile': 'INFO',
            'ComparisonFile': 'INFO',
            'SOPInstanceUID': 'INFO',
            'TagName': 'NO_DIFFERENCES_FOUND',
            'TagKeyword': 'NO_DIFFERENCES_FOUND',
            'BaselineValue': 'All instances match perfectly',
            'ComparisonValue': 'All instances match perfectly',
            'DifferenceType': 'INFO',
            'VR': 'INFO'
        })
    
    console.print(f"📊 Generated {len(rows)} report rows ({difference_count} actual differences)", style="cyan")
    
    df = pd.DataFrame(rows)
    df.to_csv(report_path, index=False)

def generate_excel_report(summary: 'ComparisonSummary', report_path: Path) -> None:
    """Generate comprehensive Excel report with charts and summary data"""
    #if not EXCEL_AVAILABLE:
    #    console.print("📊 Excel dependencies not available - generating CSV instead", style="yellow")
    #    csv_path = report_path.with_suffix('.csv')
    #    generate_csv_report(summary, csv_path)
    #    return
    
    try:
        console.print("📊 Creating Excel report with charts...", style="cyan")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create worksheets
        summary_ws = wb.create_sheet("Executive Summary")
        comparison_ws = wb.create_sheet("Comparison Results")
        tag_analysis_ws = wb.create_sheet("Tag Analysis")
        detailed_ws = wb.create_sheet("Detailed Differences")
        
        # Generate each worksheet
        _create_summary_worksheet(summary_ws, summary, wb)
        _create_comparison_worksheet(comparison_ws, summary)
        _create_tag_analysis_worksheet(tag_analysis_ws, summary)
        _create_detailed_worksheet(detailed_ws, summary)
        
        # Save workbook
        wb.save(report_path)
        console.print(f"✅ Excel report saved: {report_path}", style="green")
        
    except Exception as e:
        console.print(f"📊 Excel generation failed: {e} - generating CSV instead", style="yellow")
        csv_path = report_path.with_suffix('.csv')
        generate_csv_report(summary, csv_path)

def _auto_adjust_column_widths(ws, min_width: int = 10, max_width: int = 60) -> None:
    """
    Enhanced auto-adjust column widths based on content
    """
    # Dictionary to track the maximum content length per column
    column_widths = {}
    
    # Iterate through all rows and columns to find content
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                column_letter = cell.column_letter
                
                # Convert value to string and measure length
                cell_value = str(cell.value)
                
                # Add extra space for headers and bold text
                if cell.font and cell.font.bold:
                    content_length = len(cell_value) + 4  # Extra padding for headers
                elif cell.font and cell.font.size and cell.font.size > 12:
                    content_length = len(cell_value) + 2  # Extra padding for large text
                else:
                    content_length = len(cell_value)
                
                # Track the maximum width needed for this column
                if column_letter not in column_widths:
                    column_widths[column_letter] = content_length
                else:
                    column_widths[column_letter] = max(column_widths[column_letter], content_length)
    
    # Apply the calculated widths
    for column_letter, width in column_widths.items():
        # Apply min/max constraints
        final_width = max(min_width, min(width + 5, max_width))  # +2 for padding
        ws.column_dimensions[column_letter].width = final_width
        
        # Optional: Show what widths are being applied
        #if ws.title == "Executive Summary":  # Only show for summary sheet
        #    console.print(f"   📏 Column {column_letter}: {final_width} chars", style="dim")

def _create_summary_worksheet(ws, summary: 'ComparisonSummary', wb) -> None:
    """Create executive summary worksheet with charts and auto-sized columns"""
    # Set worksheet title
    ws.title = "Executive Summary"
    
    # Header styling
    header_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='2F5597')
    
    # Title
    ws['A1'] = "DICOM Comparison Report - Executive Summary"
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='2F5597')
    ws.merge_cells('A1:H1')
    
    # Basic information section
    ws['A3'] = "Report Information"
    ws['A3'].font = subheader_font
    
    info_data = [
        ("Baseline File:", Path(summary.baseline_file).name),
        ("Comparison Files:", f"{len(summary.comparison_files)} files"),
        ("Total Instances:", summary.total_instances),
        ("Total Studies:", summary.total_studies),
        ("Generated:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for idx, (label, value) in enumerate(info_data, 4):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    
    # Comparison Summary Table
    ws['A9'] = "Comparison Summary"
    ws['A9'].font = subheader_font
    
    # Create summary table headers
    headers = ["File Name", "Perfect Matches", "Tag Differences", "Missing Instances", "Extra Instances", "Data Integrity %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=10, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Populate summary data with better formatting
    for row_idx, result in enumerate(summary.file_results, 11):
        perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
        tag_diffs = len(result.matched_instances) - perfect_matches
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        integrity = _calculate_data_integrity(result)
        
        # File name (truncated if too long)
        file_name = Path(result.comparison_file).name
        if len(file_name) > 30:
            file_name = file_name[:27] + "..."
        
        ws.cell(row=row_idx, column=1, value=file_name)
        ws.cell(row=row_idx, column=2, value=perfect_matches)
        ws.cell(row=row_idx, column=3, value=tag_diffs)
        ws.cell(row=row_idx, column=4, value=missing)
        ws.cell(row=row_idx, column=5, value=extra)
        
        # Format integrity percentage with color coding
        integrity_cell = ws.cell(row=row_idx, column=6, value=f"{integrity:.1f}%")
        if integrity >= 95:
            integrity_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif integrity >= 85:
            integrity_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            integrity_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Add charts
    try:
        chart_start_row = len(summary.file_results) + 13
        _add_data_integrity_chart(ws, summary, start_row=chart_start_row)
        _add_comparison_breakdown_chart(ws, summary, start_row=chart_start_row, start_col=7)
    except Exception as e:
        console.print(f"⚠️  Chart creation failed: {e}", style="yellow")
    
    # Auto-adjust ALL column widths based on content
    console.print("📏 Auto-sizing columns...", style="cyan")
    _auto_adjust_column_widths(ws)

def _add_data_integrity_chart(ws, summary: 'ComparisonSummary', start_row: int) -> None:
    """Add data integrity pie chart"""
    try:
        chart = PieChart()
        chart.title = "Data Integrity Overview"
        chart.width = 15
        chart.height = 10
        
        # Calculate overall integrity
        total_perfect = 0
        total_partial = 0
        total_missing = 0
        
        for result in summary.file_results:
            perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
            tag_diffs = len(result.matched_instances) - perfect_matches
            missing = len(result.missing_instances)
            
            total_perfect += perfect_matches
            total_partial += tag_diffs
            total_missing += missing
        
        # Create data for chart
        chart_data = [
            ["Category", "Count"],
            ["Perfect Match", total_perfect],
            ["Tag Differences", total_partial],
            ["Missing Instances", total_missing]
        ]
        
        # Add data to worksheet
        chart_start_row = start_row
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=chart_start_row + row_idx, column=1 + col_idx, value=value)
        
        # Create chart reference
        data_ref = Reference(ws, min_col=2, min_row=chart_start_row + 1, max_row=chart_start_row + len(chart_data) - 1)
        labels_ref = Reference(ws, min_col=1, min_row=chart_start_row + 1, max_row=chart_start_row + len(chart_data) - 1)
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(labels_ref)
        
        # Simplified color coding (skip if it causes issues)
        try:
            if chart.series and len(chart.series) > 0:
                colors = ['00B050', 'FFC000', 'C5504B']  # Green, Orange, Red
                series = chart.series[0]
                for i, color in enumerate(colors):
                    if i < len(chart_data) - 1:  # Skip header
                        point = DataPoint(idx=i)
                        point.graphicalProperties.solidFill = color
                        series.data_points.append(point)
        except Exception as color_error:
            console.print(f"⚠️  Chart coloring skipped: {color_error}", style="dim")
        
        ws.add_chart(chart, f"A{start_row + 5}")
        
    except Exception as e:
        console.print(f"⚠️  Pie chart creation failed: {e}", style="yellow")

def _add_comparison_breakdown_chart(ws, summary: 'ComparisonSummary', start_row: int, start_col: int) -> None:
    """Add comparison breakdown bar chart"""
    try:
        chart = BarChart()
        chart.title = "File Comparison Breakdown"
        chart.x_axis.title = "Files"
        chart.y_axis.title = "Number of Instances"
        chart.width = 15
        chart.height = 10
        
        # Prepare data
        file_names = []
        perfect_matches = []
        tag_diffs = []
        missing_instances = []
        
        for result in summary.file_results:
            file_names.append(Path(result.comparison_file).name[:15])  # Truncate long names
            perfect_count = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
            perfect_matches.append(perfect_count)
            tag_diffs.append(len(result.matched_instances) - perfect_count)
            missing_instances.append(len(result.missing_instances))
        
        # Create data table with series labels in first column
        chart_data = [
            ["Series"] + ["Perfect Matches", "Tag Differences", "Missing Instances"],
        ]
        
        # Add file data as columns
        for i, file_name in enumerate(file_names):
            chart_data.append([
                file_name,
                perfect_matches[i],
                tag_diffs[i], 
                missing_instances[i]
            ])
        
        # Add data to worksheet
        chart_start_col = start_col
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=chart_start_col + col_idx, value=value)
        
        # Create chart references
        # Categories are the file names (first column, excluding header)
        categories_ref = Reference(ws, 
                                 min_col=chart_start_col, 
                                 min_row=start_row + 1, 
                                 max_row=start_row + len(file_names))
        
        # Data series are the columns (excluding first column and header)
        data_ref = Reference(ws,
                           min_col=chart_start_col + 1,
                           min_row=start_row,
                           max_col=chart_start_col + 3,  # 3 series columns
                           max_row=start_row + len(file_names))
        
        chart.add_data(data_ref, titles_from_data=True)  # Use titles from data
        chart.set_categories(categories_ref)
        
        # Add chart to worksheet
        col_letter = openpyxl.utils.get_column_letter(start_col)
        ws.add_chart(chart, f"{col_letter}{start_row + len(file_names) + 3}")
        
    except Exception as e:
        console.print(f"⚠️  Bar chart creation failed: {e}", style="yellow")

def _create_comparison_worksheet(ws, summary: 'ComparisonSummary') -> None:
    """Create detailed comparison results worksheet"""
    ws.title = "Comparison Results"
    
    # Create detailed comparison data
    data = []
    headers = ["File", "Total Instances", "Perfect Matches", "Perfect Match %", 
              "Tag Differences", "Tag Diff %", "Missing Instances", "Missing %",
              "Extra Instances", "Extra %", "Data Integrity %", "Quality Grade"]
    data.append(headers)
    
    for result in summary.file_results:
        perfect_matches = sum(1 for comp in result.matched_instances if comp.is_perfect_match)
        tag_diffs = len(result.matched_instances) - perfect_matches
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        
        total_baseline = result.total_instances_baseline
        total_comparison = result.total_instances_comparison
        
        # Calculate percentages
        perfect_pct = (perfect_matches / total_baseline * 100) if total_baseline > 0 else 0
        tag_diff_pct = (tag_diffs / total_baseline * 100) if total_baseline > 0 else 0
        missing_pct = (missing / total_baseline * 100) if total_baseline > 0 else 0
        extra_pct = (extra / total_comparison * 100) if total_comparison > 0 else 0
        integrity = _calculate_data_integrity(result)
        
        # Quality grade
        if integrity >= 95:
            grade = "A+"
        elif integrity >= 90:
            grade = "A"
        elif integrity >= 85:
            grade = "B+"
        elif integrity >= 80:
            grade = "B"
        elif integrity >= 70:
            grade = "C"
        else:
            grade = "D"
        
        row = [
            Path(result.comparison_file).name,
            total_comparison,
            perfect_matches,
            round(perfect_pct, 1),
            tag_diffs,
            round(tag_diff_pct, 1),
            missing,
            round(missing_pct, 1),
            extra,
            round(extra_pct, 1),
            round(integrity, 1),
            grade
        ]
        data.append(row)
    
    # Add data to worksheet
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            # Header formatting
            if row_idx == 0:
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            
            # Conditional formatting for quality grades
            elif col_idx == 11:  # Quality grade column
                if value in ['A+', 'A']:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                elif value in ['B+', 'B']:
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 20)
        ws.column_dimensions[column_letter].width = adjusted_width

def _create_tag_analysis_worksheet(ws, summary: 'ComparisonSummary') -> None:
    """Create tag analysis worksheet"""
    
    ws.title = "Tag Analysis"
    
    # Collect tag statistics
    tag_stats = defaultdict(lambda: {'missing': 0, 'extra': 0, 'value_diff': 0, 'type_diff': 0})
    
    for result in summary.file_results:
        for instance_comp in result.matched_instances:
            if not instance_comp.is_perfect_match:
                for tag_diff in instance_comp.tag_differences:
                    tag_name = tag_diff.tag_keyword
                    diff_type = tag_diff.difference_type.value
                    
                    if diff_type == 'MISSING_TAG':
                        tag_stats[tag_name]['missing'] += 1
                    elif diff_type == 'EXTRA_TAG':
                        tag_stats[tag_name]['extra'] += 1
                    elif diff_type == 'VALUE_DIFF':
                        tag_stats[tag_name]['value_diff'] += 1
                    elif diff_type == 'TYPE_DIFF':
                        tag_stats[tag_name]['type_diff'] += 1
    
    # Create data
    headers = ["Tag Name", "Missing Count", "Extra Count", "Value Changed", "Type Changed", "Total Affected", "Impact Level"]
    data = [headers]
    
    # Sort by total impact
    sorted_tags = sorted(tag_stats.items(), key=lambda x: sum(x[1].values()), reverse=True)
    
    for tag_name, stats in sorted_tags:
        total_affected = sum(stats.values())
        
        # Determine impact level
        if total_affected > 100:
            impact = "High"
        elif total_affected > 20:
            impact = "Medium"
        else:
            impact = "Low"
        
        row = [
            tag_name,
            stats['missing'],
            stats['extra'],
            stats['value_diff'],
            stats['type_diff'],
            total_affected,
            impact
        ]
        data.append(row)
    
    # Add to worksheet with formatting
    for row_idx, row_data in enumerate(data):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            if row_idx == 0:  # Header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
            elif col_idx == 6:  # Impact level
                if value == "High":
                    cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
                elif value == "Medium":
                    cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
                else:
                    cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 25)
        ws.column_dimensions[column_letter].width = adjusted_width

def _create_detailed_worksheet(ws, summary: 'ComparisonSummary') -> None:
    """Create detailed differences worksheet (same as CSV data)"""
    ws.title = "Detailed Differences"
    
    # Create detailed differences data (same as CSV)
    rows = []
    headers = ['ReportType', 'BaselineFile', 'ComparisonFile', 'SOPInstanceUID', 'TagName', 'TagKeyword', 'BaselineValue', 'ComparisonValue', 'DifferenceType', 'VR']
    rows.append(headers)
    
    for result in summary.file_results:
        # Add missing instances
        for missing_instance in result.missing_instances:
            rows.append([
                'MISSING_INSTANCE',
                Path(result.baseline_file).name,
                Path(result.comparison_file).name,
                missing_instance.sop_instance_uid,
                'MISSING_INSTANCE',
                'MISSING_INSTANCE',
                'EXISTS',
                'MISSING',
                'MISSING_INSTANCE',
                'INSTANCE'
            ])
        
        # Add extra instances
        for extra_instance in result.extra_instances:
            rows.append([
                'EXTRA_INSTANCE',
                Path(result.baseline_file).name,
                Path(result.comparison_file).name,
                extra_instance.sop_instance_uid,
                'EXTRA_INSTANCE',
                'EXTRA_INSTANCE',
                'MISSING',
                'EXISTS',
                'EXTRA_INSTANCE',
                'INSTANCE'
            ])
        
        # Add tag differences
        for instance_comp in result.matched_instances:
            if not instance_comp.is_perfect_match:
                for tag_diff in instance_comp.tag_differences:
                    rows.append([
                        'TAG_DIFFERENCE',
                        Path(result.baseline_file).name,
                        Path(result.comparison_file).name,
                        instance_comp.sop_instance_uid,
                        tag_diff.tag_name,
                        tag_diff.tag_keyword,
                        str(tag_diff.baseline_value) if tag_diff.baseline_value is not None else 'NULL',
                        str(tag_diff.comparison_value) if tag_diff.comparison_value is not None else 'NULL',
                        tag_diff.difference_type.value,
                        tag_diff.vr
                    ])
    
    # Add to worksheet
    for row_idx, row_data in enumerate(rows):
        for col_idx, value in enumerate(row_data):
            cell = ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
            
            if row_idx == 0:  # Header
                cell.font = Font(bold=True, color='FFFFFF')
                cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-adjust columns
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width

# Helper functions for inspect commands

def _display_search_results_brief(results: List, query: str, console: Console):
    """Display search results in brief format"""
    if not results:
        console.print(f"No results found for '{query}'", style="yellow")
        return

    console.print(f"\n🔍 Search results for '{query}' ({len(results)} matches):\n")

    for i, result in enumerate(results, 1):
        tag_info = result.tag_info
        level_color = _get_level_display_color(result.hierarchy_level)

        console.print(f"{i}. [bold]{tag_info.keyword}[/bold] ({tag_info.tag_number})")
        console.print(f"   {tag_info.name}")
        console.print(f"   Level: [{level_color}]{result.hierarchy_level}[/{level_color}] | "
                     f"Score: {result.similarity_score:.3f} | Occurrences: {result.occurrence_count}")

        if result.sample_values:
            values = ", ".join([f'"{v}"' for v in result.sample_values[:3]])
            console.print(f"   Values: {values}", style="dim")
        console.print()

def _list_available_patients(patients: dict, console: Console):
    """List available patients"""
    console.print(f"\nAvailable patients ({len(patients)}):")
    for i, (patient_id, patient) in enumerate(list(patients.items())[:10]):
        name = patient.demographics.get('PatientName', 'UNKNOWN')
        if hasattr(name, 'value'):
            name = name.value
        console.print(f"  {i+1}. {patient_id} ({name})", style="dim")

    if len(patients) > 10:
        console.print(f"  ... and {len(patients) - 10} more patients", style="dim")

def _list_available_studies(studies: dict, console: Console):
    """List available studies"""
    console.print(f"\nAvailable studies ({len(studies)}):")
    for i, (study_uid, study) in enumerate(list(studies.items())[:10]):
        desc = study.metadata.get('StudyDescription', 'UNKNOWN')
        if hasattr(desc, 'value'):
            desc = desc.value
        console.print(f"  {i+1}. {study_uid[:20]}... ({desc})", style="dim")

    if len(studies) > 10:
        console.print(f"  ... and {len(studies) - 10} more studies", style="dim")

def _list_available_series(series: dict, console: Console):
    """List available series"""
    console.print(f"\nAvailable series ({len(series)}):")
    for i, (series_uid, series) in enumerate(list(series.items())[:10]):
        desc = series.metadata.get('SeriesDescription', 'UNKNOWN')
        modality = series.metadata.get('Modality', 'UNKNOWN')
        if hasattr(desc, 'value'):
            desc = desc.value
        if hasattr(modality, 'value'):
            modality = modality.value
        console.print(f"  {i+1}. {series_uid[:20]}... ({modality} - {desc})", style="dim")

    if len(series) > 10:
        console.print(f"  ... and {len(series) - 10} more series", style="dim")

def _list_available_instances(instances: dict, console: Console, limit: int = 10):
    """List available instances"""
    console.print(f"\nAvailable instances ({len(instances)} total, showing {min(limit, len(instances))}):")
    for i, (sop_uid, instance) in enumerate(list(instances.items())[:limit]):
        instance_num = instance.metadata.get('InstanceNumber', 'UNKNOWN')
        if hasattr(instance_num, 'value'):
            instance_num = instance_num.value
        console.print(f"  {i+1}. {sop_uid[:20]}... (Instance #{instance_num})", style="dim")

def _display_patient_info(patients: dict, data, anonymize: bool, show_studies: bool, console: Console):
    """Display patient information"""

    if not patients:
        console.print("No patients found", style="yellow")
        return

    for patient_id, patient in patients.items():
        # Patient demographics table
        demo_table = Table(title=f"👤 Patient: {patient_id}")
        demo_table.add_column("Tag", style="cyan", width=20)
        demo_table.add_column("Value", style="white", width=40)

        # Display key patient tags
        patient_tags = ['PatientName', 'PatientID', 'PatientBirthDate', 'PatientSex',
                       'PatientAge', 'PatientWeight', 'PatientSize']

        for tag in patient_tags:
            tag_info = patient.demographics.get(tag)
            if tag_info:
                value = tag_info.value
                if anonymize and tag in ['PatientName', 'PatientID']:
                    value = f"ANON_{hash(str(value)) % 10000:04d}"
                demo_table.add_row(tag, str(value))

        demo_table.add_row("Total Instances", str(patient.total_instances))
        demo_table.add_row("Studies Count", str(len(patient.studies)))
        demo_table.add_row("Source Files", ", ".join(patient.file_sources))

        console.print(demo_table)

        # Show studies if requested
        if show_studies and patient.studies:
            studies_table = Table(title="📚 Studies")
            studies_table.add_column("Study Date", style="cyan")
            studies_table.add_column("Description", style="white", width=30)
            studies_table.add_column("Series", justify="right", style="yellow")
            studies_table.add_column("Instances", justify="right", style="green")

            for study_uid in patient.studies:
                study = data.studies.get(study_uid)
                if study:
                    study_date = study.metadata.get('StudyDate', 'UNKNOWN')
                    study_desc = study.metadata.get('StudyDescription', 'UNKNOWN')

                    if hasattr(study_date, 'value'):
                        study_date = study_date.value
                    if hasattr(study_desc, 'value'):
                        study_desc = study_desc.value

                    # Format date if possible
                    if study_date != 'UNKNOWN' and len(str(study_date)) == 8:
                        try:
                            date_str = str(study_date)
                            study_date = f"{date_str[:4]}-{date_str[4:6]}-{date_str[6:8]}"
                        except:
                            pass

                    studies_table.add_row(
                        str(study_date)[:10],
                        str(study_desc)[:30],
                        str(len(study.series)),
                        str(study.total_instances)
                    )

            console.print(studies_table)
        console.print()  # Space between patients

def _display_study_info(studies: dict, data, show_series: bool, console: Console):
    """Display study information"""

    if not studies:
        console.print("No studies found", style="yellow")
        return

    for study_uid, study in studies.items():
        # Study metadata table
        study_table = Table(title=f"📚 Study: {study_uid[:30]}...")
        study_table.add_column("Tag", style="cyan", width=25)
        study_table.add_column("Value", style="white", width=50)

        # Display key study tags
        study_tags = ['StudyDate', 'StudyTime', 'StudyDescription', 'StudyID',
                     'AccessionNumber', 'ReferringPhysicianName', 'InstitutionName']

        for tag in study_tags:
            tag_info = study.metadata.get(tag)
            if tag_info:
                value = str(tag_info.value)

                # Format dates and times
                if tag == 'StudyDate' and len(value) == 8:
                    try:
                        value = f"{value[:4]}-{value[4:6]}-{value[6:8]}"
                    except:
                        pass
                elif tag == 'StudyTime' and len(value) >= 6:
                    try:
                        value = f"{value[:2]}:{value[2:4]}:{value[4:6]}"
                    except:
                        pass

                study_table.add_row(tag, value[:50])

        study_table.add_row("Patient ID", study.patient_id)
        study_table.add_row("Series Count", str(len(study.series)))
        study_table.add_row("Total Instances", str(study.total_instances))
        study_table.add_row("Source Files", ", ".join(study.file_sources))

        console.print(study_table)

        # Show series if requested
        if show_series and study.series:
            series_table = Table(title="🔬 Series")
            series_table.add_column("Series #", style="cyan")
            series_table.add_column("Modality", style="blue")
            series_table.add_column("Description", style="white", width=30)
            series_table.add_column("Instances", justify="right", style="green")

            for series_uid in study.series:
                series = data.series.get(series_uid)
                if series:
                    series_num = series.metadata.get('SeriesNumber', 'UNKNOWN')
                    modality = series.metadata.get('Modality', 'UNKNOWN')
                    series_desc = series.metadata.get('SeriesDescription', 'UNKNOWN')

                    if hasattr(series_num, 'value'):
                        series_num = series_num.value
                    if hasattr(modality, 'value'):
                        modality = modality.value
                    if hasattr(series_desc, 'value'):
                        series_desc = series_desc.value

                    series_table.add_row(
                        str(series_num),
                        str(modality),
                        str(series_desc)[:30],
                        str(len(series.instances))
                    )

            console.print(series_table)
        console.print()

def _display_series_info(series: dict, data, show_instances: bool, console: Console):
    """Display series information"""

    if not series:
        console.print("No series found", style="yellow")
        return

    for series_uid, series in series.items():
        # Series metadata table
        series_table = Table(title=f"🔬 Series: {series_uid[:30]}...")
        series_table.add_column("Tag", style="cyan", width=25)
        series_table.add_column("Value", style="white", width=50)

        # Display key series tags
        series_tags = ['SeriesNumber', 'SeriesDate', 'SeriesTime', 'SeriesDescription',
                      'Modality', 'ProtocolName', 'BodyPartExamined', 'PatientPosition']

        for tag in series_tags:
            tag_info = series.metadata.get(tag)
            if tag_info:
                value = str(tag_info.value)

                # Format dates and times
                if tag == 'SeriesDate' and len(value) == 8:
                    try:
                        value = f"{value[:4]}-{value[4:6]}-{value[6:8]}"
                    except:
                        pass
                elif tag == 'SeriesTime' and len(value) >= 6:
                    try:
                        value = f"{value[:2]}:{value[2:4]}:{value[4:6]}"
                    except:
                        pass

                series_table.add_row(tag, value[:50])

        series_table.add_row("Study UID", series.study_uid[:30] + "...")
        series_table.add_row("Instance Count", str(len(series.instances)))
        series_table.add_row("Source Files", ", ".join(series.file_sources))

        console.print(series_table)

        # Show instances if requested
        if show_instances and series.instances:
            instances_table = Table(title="🖼️  Instances")
            instances_table.add_column("Instance #", style="cyan")
            instances_table.add_column("SOP Class", style="blue", width=25)
            instances_table.add_column("File Path", style="white", width=40)

            for sop_uid in series.instances[:10]:  # Limit to first 10
                instance = data.instances.get(sop_uid)
                if instance:
                    instance_num = instance.metadata.get('InstanceNumber', 'UNKNOWN')
                    sop_class = instance.metadata.get('SOPClassUID', 'UNKNOWN')

                    if hasattr(instance_num, 'value'):
                        instance_num = instance_num.value
                    if hasattr(sop_class, 'value'):
                        sop_class = sop_class.value

                    # Shorten SOP class for display
                    if len(str(sop_class)) > 25:
                        sop_class = str(sop_class)[-25:] + "..."

                    instances_table.add_row(
                        str(instance_num),
                        str(sop_class)[:25],
                        str(instance.file_path)[:40]
                    )

            if len(series.instances) > 10:
                instances_table.add_row("...", f"and {len(series.instances) - 10} more", "...")

            console.print(instances_table)
        console.print()

def _display_instance_info(instances: dict, data, show_all_tags: bool, console: Console):
    """Display instance information"""

    if not instances:
        console.print("No instances found", style="yellow")
        return

    for sop_uid, instance in instances.items():
        # Instance metadata table
        instance_table = Table(title=f"🖼️  Instance: {sop_uid[:30]}...")
        instance_table.add_column("Tag", style="cyan", width=25)
        instance_table.add_column("Value", style="white", width=50)

        if show_all_tags:
            # Show all tags
            for tag_keyword, tag_info in instance.metadata.items():
                value = str(tag_info.value)
                instance_table.add_row(f"{tag_keyword} ({tag_info.tag_number})", value[:50])
        else:
            # Display key instance tags
            instance_tags = ['InstanceNumber', 'SOPClassUID', 'SOPInstanceUID',
                           'ImageType', 'Rows', 'Columns', 'BitsAllocated',
                           'PhotometricInterpretation', 'SliceLocation', 'SliceThickness']

            for tag in instance_tags:
                tag_info = instance.metadata.get(tag)
                if tag_info:
                    value = str(tag_info.value)
                    instance_table.add_row(tag, value[:50])

        instance_table.add_row("Series UID", instance.series_uid[:30] + "...")
        instance_table.add_row("Source File", instance.source_file)
        instance_table.add_row("File Path", str(instance.file_path))

        console.print(instance_table)
        console.print()

def _get_level_display_color(level: str) -> str:
    """Get color for hierarchy level display"""
    colors = {
        'patient': 'green',
        'study': 'blue',
        'series': 'yellow',
        'instance': 'magenta'
    }
    return colors.get(level, 'white')

if __name__ == "__main__":
    app()