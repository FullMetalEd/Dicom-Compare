"""Image comparison command implementation"""

from typing import List, Optional
from pathlib import Path
import typer
from rich.console import Console
from rich.panel import Panel
from rich.table import Table
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.chart import PieChart, BarChart, ScatterChart, Reference
from openpyxl.chart.series import DataPoint
import pandas as pd
import numpy as np

from .image_comparator import ImageComparator
from .image_models import ImageComparisonSummary, ImageFileComparisonResult
from .dicom_extractor import DicomExtractor
from .dicom_loader import DicomLoader
from .utils import validate_inputs, create_temp_dir, cleanup_temp_dirs

console = Console()

def run_image_comparison(
    files: List[Path],
    report: Optional[Path] = None,
    tolerance: float = 0.0,
    normalize: bool = True,
    verbose: bool = False
) -> None:
    """Main image comparison workflow"""
    
    console.print("🖼️ Starting DICOM image comparison...", style="blue")
    
    # Validate inputs (reuse existing validation)
    try:
        validate_inputs(files)
        if report:
            validate_image_report_path(report)
    except Exception as e:
        console.print(f"❌ {str(e)}", style="red")
        raise typer.Exit(1)
    
    temp_dirs = []
    extraction_stats = []
    try:
        # Extract ZIP files
        console.print("📦 Extracting ZIP files...", style="yellow")
        extractor = DicomExtractor(verbose=verbose)
        extracted_paths = []
        
        for file in files:
            temp_dir = create_temp_dir()
            temp_dirs.append(temp_dir)
            extracted_path, stats = extractor.extract_zip(file, temp_dir)
            extracted_paths.append((str(file), extracted_path))
            extraction_stats.append((str(file), stats))
        
        # Load DICOM files
        console.print("🏥 Loading DICOM files...", style="yellow")
        loader = DicomLoader(verbose=verbose)
        loaded_studies = []
        
        for i, (file_name, path) in enumerate(extracted_paths):
            studies = loader.load_dicom_files(path, file_name)
            loaded_studies.append((file_name, studies))
            
            # Show results with extraction context
            total_instances = sum(len(series.instances) for study in studies.values() 
                                for series in study.series.values())
            
            # Get corresponding extraction stats
            _, stats = extraction_stats[i]
            
            console.print(f"   {Path(file_name).name}: {total_instances} instances", style="cyan")
            
            if verbose:
                console.print(f"     Found {stats.dicom_files} DICOM files with pixel data", style="dim")
        
        # Image comparison
        console.print("🔍 Comparing image pixel data...", style="yellow")
        image_comparator = ImageComparator(tolerance=tolerance, normalize=normalize)
        
        baseline_name, baseline_studies = loaded_studies[0]
        comparison_results = []
        
        for comp_name, comp_studies in loaded_studies[1:]:
            result = image_comparator.compare_studies(
                baseline_studies, comp_studies,
                baseline_name, comp_name
            )
            comparison_results.append(result)
        
        # Create summary
        summary = create_image_comparison_summary(baseline_name, comparison_results, tolerance, normalize)
        
        # Display results
        display_image_terminal_results(summary, console)
        
        # Generate report if requested
        if report:
            console.print(f"📊 Generating image report: {report}", style="green")
            generate_image_report(summary, report)
            console.print(f"✅ Report saved to: {report}", style="green")
        
        console.print("🎉 Image comparison completed successfully!", style="green")
        
    except Exception as e:
        console.print(f"❌ Error during image comparison: {str(e)}", style="red")
        if verbose:
            console.print_exception()
        raise typer.Exit(1)
    
    finally:
        cleanup_temp_dirs(temp_dirs)

def validate_image_report_path(report_path: Path) -> None:
    """Validate image report path and format"""
    if not report_path.suffix.lower() in ['.csv', '.xlsx']:
        raise ValueError("Image report format must be CSV (.csv) or Excel (.xlsx)")
    
    report_path.parent.mkdir(parents=True, exist_ok=True)

def create_image_comparison_summary(
    baseline_name: str, 
    results: List[ImageFileComparisonResult],
    tolerance: float,
    normalize: bool
) -> ImageComparisonSummary:
    """Create summary from image comparison results"""
    comparison_files = [result.comparison_file for result in results]
    total_images = sum(len(result.image_comparisons) for result in results)
    
    return ImageComparisonSummary(
        baseline_file=baseline_name,
        comparison_files=comparison_files,
        file_results=results,
        tolerance_used=tolerance,
        normalization_applied=normalize,
        total_images_compared=total_images
    )

def display_image_terminal_results(summary: ImageComparisonSummary, console: Console) -> None:
    """Display image comparison results in terminal"""
    
    # Create summary panel
    summary_text = f"Baseline: {Path(summary.baseline_file).name}\n"
    summary_text += f"Comparison Mode: Image Pixel Data\n"
    summary_text += f"Tolerance: {summary.tolerance_used}\n"
    summary_text += f"Normalization: {'Applied' if summary.normalization_applied else 'Disabled'}\n"
    summary_text += f"Images Compared: {summary.total_images_compared}"
    
    console.print(Panel(summary_text, title="🖼️ DICOM Image Comparison Summary", expand=False))
    
    # Results table
    table = Table(title="🖼️ Image Comparison Results")
    table.add_column("File", style="cyan", width=25)
    table.add_column("Exact\nMatches", style="green", justify="right")
    table.add_column("Pixel\nDifferences", style="yellow", justify="right")
    table.add_column("Avg\nSimilarity", style="bright_blue", justify="right")
    table.add_column("Missing\nImages", style="red", justify="right")
    table.add_column("Extra\nImages", style="magenta", justify="right")
    table.add_column("Match\n%", style="bright_green", justify="right")
    
    for result in summary.file_results:
        exact_matches = result.exact_matches
        pixel_diffs = result.pixel_differences
        avg_similarity = result.average_similarity
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        
        # Calculate match percentage
        total_baseline = result.total_instances_baseline
        match_pct = (exact_matches / total_baseline * 100) if total_baseline > 0 else 0
        
        table.add_row(
            Path(result.comparison_file).name,
            str(exact_matches),
            str(pixel_diffs),
            f"{avg_similarity:.1%}",
            str(missing),
            str(extra),
            f"{match_pct:.1f}%"
        )
    
    console.print(table)
    
    # Add detailed statistics if there are differences
    _display_image_statistics(summary, console)

def _display_image_statistics(summary: ImageComparisonSummary, console: Console) -> None:
    """Display detailed image comparison statistics"""
    
    # Collect statistics
    total_comparisons = 0
    total_exact = 0
    total_differences = 0
    max_difference_found = 0.0
    
    for result in summary.file_results:
        total_comparisons += len(result.image_comparisons)
        total_exact += result.exact_matches
        total_differences += result.pixel_differences
        
        for img_comp in result.image_comparisons:
            if img_comp.max_difference is not None:
                max_difference_found = max(max_difference_found, img_comp.max_difference)
    
    if total_comparisons > 0 and total_differences > 0:
        console.print("\n")
        
        stats_table = Table(title="📈 Image Difference Statistics")
        stats_table.add_column("Metric", style="cyan")
        stats_table.add_column("Value", style="bright_white", justify="right")
        
        stats_table.add_row("Total Images Compared", str(total_comparisons))
        stats_table.add_row("Exact Matches", str(total_exact))
        stats_table.add_row("Images with Differences", str(total_differences))
        stats_table.add_row("Overall Match Rate", f"{(total_exact/total_comparisons)*100:.1f}%")
        stats_table.add_row("Max Pixel Difference Found", f"{max_difference_found:.2f}")
        
        console.print(stats_table)

def generate_image_report(summary: ImageComparisonSummary, report_path: Path) -> None:
    """Generate image comparison report"""
    if report_path.suffix.lower() == '.csv':
        generate_image_csv_report(summary, report_path)
    elif report_path.suffix.lower() == '.xlsx':
        generate_image_excel_report(summary, report_path)

def generate_image_csv_report(summary: ImageComparisonSummary, report_path: Path) -> None:
    """Generate CSV report for image comparisons"""
    import pandas as pd
    
    rows = []
    headers = ['BaselineFile', 'ComparisonFile', 'SOPInstanceUID', 'ExactMatch', 
              'SimilarityScore', 'PixelDifferences', 'MaxDifference', 'MeanDifference',
              'RMSE', 'BaselineShape', 'ComparisonShape', 'DifferenceType', 'ToleranceUsed']
    
    for result in summary.file_results:
        for img_comp in result.image_comparisons:
            rows.append([
                Path(result.baseline_file).name,
                Path(result.comparison_file).name,
                img_comp.sop_instance_uid,
                img_comp.is_exact_match,
                f"{img_comp.similarity_score:.4f}",
                img_comp.pixel_differences,
                img_comp.max_difference,
                img_comp.mean_difference,
                img_comp.rmse,
                str(img_comp.baseline_stats.shape) if img_comp.baseline_stats else "N/A",
                str(img_comp.comparison_stats.shape) if img_comp.comparison_stats else "N/A",
                img_comp.difference_type.value,
                img_comp.tolerance_used
            ])
    
    df = pd.DataFrame(rows, columns=headers)
    df.to_csv(report_path, index=False)
    
    console.print(f"📊 Generated CSV with {len(rows)} image comparisons", style="cyan")

def generate_image_excel_report(summary: ImageComparisonSummary, report_path: Path) -> None:
    """Generate comprehensive Excel report for image comparisons"""   
    try:
        console.print("📊 Creating Excel image comparison report...", style="cyan")
        
        # Create workbook
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Create worksheets
        summary_ws = wb.create_sheet("Image Summary")
        comparison_ws = wb.create_sheet("Detailed Comparisons")
        statistics_ws = wb.create_sheet("Statistics")
        settings_ws = wb.create_sheet("Settings & Info")
        
        # Generate each worksheet
        _create_image_summary_worksheet(summary_ws, summary)
        _create_image_comparison_worksheet(comparison_ws, summary)
        _create_image_statistics_worksheet(statistics_ws, summary)
        _create_image_settings_worksheet(settings_ws, summary)
        
        # Save workbook
        wb.save(report_path)
        console.print(f"✅ Excel image report saved: {report_path}", style="green")
        
    except ImportError:
        console.print("📊 Excel dependencies not available - generating CSV instead", style="yellow")
        csv_path = report_path.with_suffix('.csv')
        generate_image_csv_report(summary, csv_path)
    except Exception as e:
        console.print(f"📊 Excel generation failed: {e} - generating CSV instead", style="yellow")
        csv_path = report_path.with_suffix('.csv')
        generate_image_csv_report(summary, csv_path)

def _create_image_summary_worksheet(ws, summary: ImageComparisonSummary) -> None:
    """Create image comparison summary worksheet with charts"""
    ws.title = "Image Summary"
    
    # Styling
    header_font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='2F5597')
    
    # Title
    ws['A1'] = "DICOM Image Comparison Report"
    ws['A1'].font = Font(name='Calibri', size=18, bold=True, color='2F5597')
    ws.merge_cells('A1:H1')
    
    # Summary information
    ws['A3'] = "Report Information"
    ws['A3'].font = subheader_font
    
    info_data = [
        ("Baseline File:", Path(summary.baseline_file).name),
        ("Comparison Files:", f"{len(summary.comparison_files)} files"),
        ("Total Images Compared:", summary.total_images_compared),
        ("Tolerance Used:", summary.tolerance_used),
        ("Normalization Applied:", "Yes" if summary.normalization_applied else "No"),
        ("Overall Similarity:", f"{summary.overall_similarity:.1%}"),
        ("Generated:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S"))
    ]
    
    for idx, (label, value) in enumerate(info_data, 4):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    
    # Results summary table
    ws['A12'] = "Comparison Results Summary"
    ws['A12'].font = subheader_font
    
    # Create summary table
    headers = ["File", "Total Images", "Exact Matches", "Pixel Differences", 
              "Avg Similarity", "Missing Images", "Extra Images", "Match %"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=13, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
    
    # Populate data
    for row_idx, result in enumerate(summary.file_results, 14):
        exact_matches = result.exact_matches
        pixel_diffs = result.pixel_differences
        total_images = len(result.image_comparisons)
        avg_similarity = result.average_similarity
        missing = len(result.missing_instances)
        extra = len(result.extra_instances)
        match_pct = (exact_matches / total_images * 100) if total_images > 0 else 0
        
        ws.cell(row=row_idx, column=1, value=Path(result.comparison_file).name)
        ws.cell(row=row_idx, column=2, value=total_images)
        ws.cell(row=row_idx, column=3, value=exact_matches)
        ws.cell(row=row_idx, column=4, value=pixel_diffs)
        ws.cell(row=row_idx, column=5, value=f"{avg_similarity:.1%}")
        ws.cell(row=row_idx, column=6, value=missing)
        ws.cell(row=row_idx, column=7, value=extra)
        
        # Color-code match percentage
        match_cell = ws.cell(row=row_idx, column=8, value=f"{match_pct:.1f}%")
        if match_pct >= 95:
            match_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
        elif match_pct >= 85:
            match_cell.fill = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
        else:
            match_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
    
    # Add charts
    chart_start_row = len(summary.file_results) + 16
    _add_image_similarity_chart(ws, summary, chart_start_row)
    _add_image_difference_breakdown_chart(ws, summary, chart_start_row, start_col=7)
    
    # Auto-adjust columns
    _auto_adjust_column_widths(ws)

def _add_image_similarity_chart(ws, summary: ImageComparisonSummary, start_row: int) -> None:
    """Add image similarity pie chart (safer version)"""
    try:
        chart = PieChart()
        chart.title = "Image Similarity Overview"
        chart.width = 15
        chart.height = 10
        
        # Calculate totals
        total_exact = summary.overall_exact_matches
        total_differences = sum(result.pixel_differences for result in summary.file_results)
        total_missing = sum(len(result.missing_instances) for result in summary.file_results)
        
        # Only create chart if we have meaningful data
        if total_exact + total_differences + total_missing == 0:
            console.print("⚠️  No data for similarity chart", style="yellow")
            return
        
        # Create chart data
        chart_data = [
            ["Category", "Count"],
            ["Exact Matches", total_exact],
            ["Pixel Differences", total_differences],
            ["Missing Images", total_missing]
        ]
        
        # Add data to worksheet
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=1 + col_idx, value=value)
        
        # Create chart with proper error handling
        try:
            data_ref = Reference(ws, min_col=2, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
            labels_ref = Reference(ws, min_col=1, min_row=start_row + 1, max_row=start_row + len(chart_data) - 1)
            
            chart.add_data(data_ref, titles_from_data=False)
            chart.set_categories(labels_ref)
            
            # Simplified color coding (skip if it causes issues)
            try:
                if chart.series and len(chart.series) > 0:
                    series = chart.series[0]
                    colors = ['00B050', 'FFC000', 'C5504B']  # Green, Orange, Red
                    for i, color in enumerate(colors):
                        if i < 3:  # Only first 3 data points
                            point = DataPoint(idx=i)
                            point.graphicalProperties.solidFill = color
                            series.data_points.append(point)
            except Exception as color_error:
                console.print(f"⚠️  Chart coloring skipped: {color_error}", style="dim")
            
            ws.add_chart(chart, f"A{start_row + 5}")
            
        except Exception as chart_error:
            console.print(f"⚠️  Chart creation failed: {chart_error}", style="yellow")
        
    except Exception as e:
        console.print(f"⚠️  Image similarity chart failed: {e}", style="yellow")


def _add_image_difference_breakdown_chart(ws, summary: ImageComparisonSummary, start_row: int, start_col: int) -> None:
    """Add bar chart showing difference breakdown by file (fixed version)"""
    try:
        chart = BarChart()
        chart.title = "Image Differences by File"
        chart.x_axis.title = "Files"
        chart.y_axis.title = "Number of Images"
        chart.width = 15
        chart.height = 10
        
        # Prepare data
        file_names = []
        exact_matches = []
        differences = []
        
        for result in summary.file_results:
            file_names.append(Path(result.comparison_file).name[:15])  # Truncate long names
            exact_matches.append(result.exact_matches)
            differences.append(result.pixel_differences)
        
        # Create simplified data table (avoid complex series labeling)
        chart_data = [
            ["File"] + file_names,
            ["Exact Matches"] + exact_matches,
        ]
        
        # Add data to worksheet
        chart_start_col = start_col
        for row_idx, row_data in enumerate(chart_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=start_row + row_idx, column=chart_start_col + col_idx, value=value)
        
        # Create simple chart without complex series titles
        categories_ref = Reference(ws, 
                                 min_col=chart_start_col + 1, 
                                 min_row=start_row, 
                                 max_col=chart_start_col + len(file_names))
        
        data_ref = Reference(ws,
                           min_col=chart_start_col + 1,
                           min_row=start_row + 1,
                           max_col=chart_start_col + len(file_names))
        
        chart.add_data(data_ref, titles_from_data=False)
        chart.set_categories(categories_ref)
        
        # Add chart to worksheet
        col_letter = openpyxl.utils.get_column_letter(start_col)
        ws.add_chart(chart, f"{col_letter}{start_row + 4}")
        
    except Exception as e:
        console.print(f"⚠️  Image breakdown chart failed: {e}", style="yellow")


def _create_image_comparison_worksheet(ws, summary: ImageComparisonSummary) -> None:
    """Create detailed image comparison worksheet"""
    ws.title = "Detailed Comparisons"
    
    # Headers
    headers = ["Baseline File", "Comparison File", "SOP Instance UID", "Exact Match", 
              "Similarity Score", "Pixel Differences", "Max Difference", "Mean Difference",
              "RMSE", "Baseline Shape", "Comparison Shape", "Difference Type", "Tolerance Used"]
    
    # Add headers with formatting
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
        cell.alignment = Alignment(horizontal='center')
    
    # Add data
    row_idx = 2
    for result in summary.file_results:
        for img_comp in result.image_comparisons:
            ws.cell(row=row_idx, column=1, value=Path(result.baseline_file).name)
            ws.cell(row=row_idx, column=2, value=Path(result.comparison_file).name)
            ws.cell(row=row_idx, column=3, value=img_comp.sop_instance_uid)
            
            # Color-code exact match
            exact_match_cell = ws.cell(row=row_idx, column=4, value=img_comp.is_exact_match)
            if img_comp.is_exact_match:
                exact_match_cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
            else:
                exact_match_cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            
            ws.cell(row=row_idx, column=5, value=f"{img_comp.similarity_score:.4f}")
            ws.cell(row=row_idx, column=6, value=img_comp.pixel_differences)
            ws.cell(row=row_idx, column=7, value=img_comp.max_difference)
            ws.cell(row=row_idx, column=8, value=img_comp.mean_difference)
            ws.cell(row=row_idx, column=9, value=img_comp.rmse)
            ws.cell(row=row_idx, column=10, value=str(img_comp.baseline_stats.shape) if img_comp.baseline_stats else "N/A")
            ws.cell(row=row_idx, column=11, value=str(img_comp.comparison_stats.shape) if img_comp.comparison_stats else "N/A")
            ws.cell(row=row_idx, column=12, value=img_comp.difference_type.value)
            ws.cell(row=row_idx, column=13, value=img_comp.tolerance_used)
            
            row_idx += 1
    
    # Auto-adjust columns
    _auto_adjust_column_widths(ws)

def _create_image_statistics_worksheet(ws, summary: ImageComparisonSummary) -> None:
    """Create image statistics worksheet"""
    ws.title = "Statistics"
    
    subheader_font = Font(name='Calibri', size=12, bold=True, color='2F5597')
    
    # Overall statistics
    ws['A1'] = "Image Comparison Statistics"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5597')
    
    ws['A3'] = "Overall Results"
    ws['A3'].font = subheader_font
    
    # Calculate statistics
    total_comparisons = sum(len(result.image_comparisons) for result in summary.file_results)
    total_exact = summary.overall_exact_matches
    total_differences = total_comparisons - total_exact
    
    # Collect similarity scores
    all_similarities = []
    all_max_diffs = []
    all_mean_diffs = []
    
    for result in summary.file_results:
        for img_comp in result.image_comparisons:
            all_similarities.append(img_comp.similarity_score)
            if img_comp.max_difference is not None:
                all_max_diffs.append(img_comp.max_difference)
            if img_comp.mean_difference is not None:
                all_mean_diffs.append(img_comp.mean_difference)
    
    # Statistics table
    stats_data = [
        ("Total Images Compared:", total_comparisons),
        ("Exact Matches:", total_exact),
        ("Images with Differences:", total_differences),
        ("Overall Match Rate:", f"{(total_exact/total_comparisons)*100:.1f}%" if total_comparisons > 0 else "0%"),
        ("Average Similarity:", f"{np.mean(all_similarities):.1%}" if all_similarities else "N/A"),
        ("Min Similarity:", f"{np.min(all_similarities):.1%}" if all_similarities else "N/A"),
        ("Max Pixel Difference:", f"{np.max(all_max_diffs):.2f}" if all_max_diffs else "N/A"),
        ("Average Pixel Difference:", f"{np.mean(all_mean_diffs):.2f}" if all_mean_diffs else "N/A"),
    ]
    
    for idx, (label, value) in enumerate(stats_data, 4):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    
    # Per-file breakdown
    ws['A14'] = "Per-File Statistics"
    ws['A14'].font = subheader_font
    
    file_headers = ["File", "Images", "Exact Matches", "Differences", "Match %", "Avg Similarity"]
    for col, header in enumerate(file_headers, 1):
        cell = ws.cell(row=15, column=col, value=header)
        cell.font = Font(bold=True)
    
    for row_idx, result in enumerate(summary.file_results, 16):
        total_images = len(result.image_comparisons)
        exact_matches = result.exact_matches
        differences = result.pixel_differences
        match_pct = (exact_matches / total_images * 100) if total_images > 0 else 0
        avg_similarity = result.average_similarity
        
        ws.cell(row=row_idx, column=1, value=Path(result.comparison_file).name)
        ws.cell(row=row_idx, column=2, value=total_images)
        ws.cell(row=row_idx, column=3, value=exact_matches)
        ws.cell(row=row_idx, column=4, value=differences)
        ws.cell(row=row_idx, column=5, value=f"{match_pct:.1f}%")
        ws.cell(row=row_idx, column=6, value=f"{avg_similarity:.1%}")
    
    _auto_adjust_column_widths(ws)

def _create_image_settings_worksheet(ws, summary: ImageComparisonSummary) -> None:
    """Create settings and information worksheet"""
    ws.title = "Settings & Info"
    
    subheader_font = Font(name='Calibri', size=12, bold=True, color='2F5597')
    
    # Settings used
    ws['A1'] = "Comparison Settings"
    ws['A1'].font = Font(name='Calibri', size=16, bold=True, color='2F5597')
    
    settings_data = [
        ("Tolerance Used:", summary.tolerance_used),
        ("Normalization Applied:", "Yes" if summary.normalization_applied else "No"),
        ("Comparison Mode:", "Image Pixel Data"),
        ("Report Generated:", pd.Timestamp.now().strftime("%Y-%m-%d %H:%M:%S")),
    ]
    
    for idx, (label, value) in enumerate(settings_data, 3):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=value)
    
    # File information
    ws['A8'] = "File Information"
    ws['A8'].font = subheader_font
    
    ws.cell(row=9, column=1, value="Baseline File:").font = Font(bold=True)
    ws.cell(row=9, column=2, value=Path(summary.baseline_file).name)
    
    ws.cell(row=11, column=1, value="Comparison Files:").font = Font(bold=True)
    for idx, comp_file in enumerate(summary.comparison_files, 12):
        ws.cell(row=idx, column=2, value=Path(comp_file).name)
    
    # Explanation of metrics
    ws[f'A{15 + len(summary.comparison_files)}'] = "Metric Explanations"
    ws[f'A{15 + len(summary.comparison_files)}'].font = subheader_font
    
    explanations = [
        ("Similarity Score:", "Percentage of pixels that match (0.0 to 1.0)"),
        ("RMSE:", "Root Mean Square Error - lower values indicate more similar images"),
        ("Max Difference:", "Largest pixel value difference found between images"),
        ("Mean Difference:", "Average pixel value difference across all pixels"),
        ("Tolerance:", "Maximum allowed pixel difference to be considered a match"),
        ("Normalization:", "Applies DICOM rescale slope/intercept and windowing"),
    ]
    
    start_row = 17 + len(summary.comparison_files)
    for idx, (term, explanation) in enumerate(explanations):
        ws.cell(row=start_row + idx, column=1, value=term).font = Font(bold=True)
        ws.cell(row=start_row + idx, column=2, value=explanation)
    
    _auto_adjust_column_widths(ws)

def _auto_adjust_column_widths(ws, min_width: int = 10, max_width: int = 50) -> None:
    """Auto-adjust column widths based on content (fixed for merged cells)"""
    column_widths = {}
    
    # Iterate through all rows and columns safely
    for row in ws.iter_rows():
        for cell in row:
            # Skip merged cells
            if hasattr(cell, 'coordinate') and cell.coordinate in ws.merged_cells:
                continue
            
            if cell.value is not None:
                # Get column letter safely
                try:
                    column_letter = cell.column_letter
                except AttributeError:
                    # Handle merged cells or other edge cases
                    continue
                
                # Calculate content length
                cell_value = str(cell.value)
                content_length = len(cell_value)
                
                # Add extra space for headers and bold text
                if cell.font and cell.font.bold:
                    content_length += 4
                elif cell.font and cell.font.size and cell.font.size > 12:
                    content_length += 2
                
                # Track the maximum width needed for this column
                if column_letter not in column_widths:
                    column_widths[column_letter] = content_length
                else:
                    column_widths[column_letter] = max(column_widths[column_letter], content_length)
    
    # Apply the calculated widths
    for column_letter, width in column_widths.items():
        final_width = max(min_width, min(width + 2, max_width))
        ws.column_dimensions[column_letter].width = final_width